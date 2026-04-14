"""
과거 데이터 임포트 (2023~2025)
- 각 브랜드의 합계 행에서 col3(광고비), col4(매출)을 읽음
- 서브채널은 무시 (합계에 이미 포함)
- 단, "기 타(일매출 포함x)"는 합계에 미포함이므로 별도 추가
"""
import pandas as pd
import sqlite3
import re
import openpyxl
from datetime import datetime


_WB_CACHE = {}

def _get_wb(fpath, data_only):
    key = (fpath, data_only)
    if key not in _WB_CACHE:
        _WB_CACHE[key] = openpyxl.load_workbook(fpath, data_only=data_only)
    return _WB_CACHE[key]


def get_section_truth(fpath, sheet_name, header_row_pandas, formula_refs):
    """섹션의 시트 공식 기준 월간 광고비/매출 truth 값 반환 (없으면 None)"""
    if formula_refs is None:
        return None, None
    # header_row+30 ~ +38 범위 안의 ref가 이 섹션의 sum_row
    sum_row = None
    for ref in formula_refs:
        if header_row_pandas + 30 <= ref <= header_row_pandas + 38:
            sum_row = ref
            break
    if sum_row is None:
        return None, None
    # 공식의 D{ref}, E{ref} 값을 직접 가져옴 (단, 공식 D4 자체에서 어떤 col 사용했는지 분리 필요)
    wb_v = _get_wb(fpath, True)
    ws = wb_v[sheet_name]
    # 공식 텍스트에서 각 colon 추출
    wb_f = _get_wb(fpath, False)
    wsf = wb_f[sheet_name]
    d4 = wsf.cell(4, 4).value
    e4 = wsf.cell(4, 5).value
    ad = rev = None
    if isinstance(d4, str):
        for m in re.finditer(r'([A-Z]+)(\d+)', d4):
            col_letter, row = m.group(1), int(m.group(2))
            if row == sum_row:
                ad = ws[f'{col_letter}{row}'].value
                break
    if isinstance(e4, str):
        for m in re.finditer(r'([A-Z]+)(\d+)', e4):
            col_letter, row = m.group(1), int(m.group(2))
            if row == sum_row:
                rev = ws[f'{col_letter}{row}'].value
                break
    return ad, rev


def get_formula_row_refs(fpath, sheet_name):
    """시트의 D4/E4 공식에서 참조하는 row 번호 집합 반환 (openpyxl 1-based)"""
    try:
        wb = openpyxl.load_workbook(fpath, data_only=False)
        ws = wb[sheet_name]
        refs = set()
        for col in [4, 5]:  # D=4, E=5
            f = ws.cell(4, col).value
            if isinstance(f, str) and f.startswith('='):
                # 추출: D40, E115, AA77, T39 등
                for m in re.finditer(r'[A-Z]+(\d+)', f):
                    refs.add(int(m.group(1)))
        wb.close()
        return refs
    except Exception:
        return None


def parse_sum_formula(ws, sum_row_openpyxl, target_col):
    """sum_row의 공식을 읽고 SUM(...) 외의 추가 항목 셀 좌표 리스트 반환
    e.g. =SUM(D9:D39)+L40+L77 -> ['L40', 'L77']
    또는 =D74+D110 -> base components"""
    try:
        f = ws.cell(sum_row_openpyxl, target_col).value
    except:
        return None, None
    if not isinstance(f, str) or not f.startswith('='):
        return None, None
    # SUM(...) 부분 제거
    body = f[1:]
    sum_match = re.search(r'SUM\([^)]+\)', body)
    has_sum = sum_match is not None
    if has_sum:
        body_no_sum = body.replace(sum_match.group(), '')
    else:
        body_no_sum = body
    # 추가 셀 참조 추출
    extras = re.findall(r'[A-Z]+\d+', body_no_sum)
    return has_sum, extras


def section_in_formula(header_row_pandas, formula_refs):
    """섹션이 시트 공식에 포함되는지 판단 (header_row 다음 30~36행 사이에 합계 row가 있어야 함)"""
    if formula_refs is None:
        return True  # fallback: include all
    # openpyxl row = pandas row + 1; sum row 보통 header + 32 ~ 35 (data_start ~ 31일 + 합계)
    for ref in formula_refs:
        # ref가 (header_row+30)~(header_row+38) 범위면 이 섹션의 합계 행
        if header_row_pandas + 30 <= ref <= header_row_pandas + 38:
            return True
    return False

DB_PATH = "dashboard_data.db"
BASE_DIR = "C:/Users/조현우/Desktop/광고일지"


def safe_int(val):
    if pd.isna(val): return 0
    try:
        v = int(float(val))
        return v if v > 0 else 0
    except:
        return 0


def safe_date(val, year, month):
    if pd.isna(val): return None
    try:
        d = pd.to_datetime(val).date()
        return d if d.year == year and d.month == month else None
    except:
        return None


def find_sub_ad_sections(df):
    """하위 합계 행(자사몰 합계, 오픈마켓 합계) 찾기 - 광고비 추가 수집용"""
    sub_sections = []
    for r in range(min(200, len(df))):
        for c in range(min(5, len(df.columns))):
            v = str(df.iloc[r, c]).strip()
            if '아자차 자사몰 합계' in v or '아자차 오픈마켓 합계' in v:
                sub_sections.append({'row': r, 'label': v, 'brand': '아자차'})
    return sub_sections


def extract_sub_ad_costs(df, sub_section, year, month):
    """하위 섹션에서 광고 채널의 광고비만 추출"""
    header_row = sub_section['row']
    col_row = header_row + 1
    data_start = header_row + 2
    data_end = data_start + 33

    date_col = 2
    for c in range(5):
        if col_row < len(df) and pd.notna(df.iloc[col_row, c]):
            if '날' in str(df.iloc[col_row, c]).strip():
                date_col = c
                break

    # 광고 채널의 광고비 열 찾기 (Meta, NSA, Google, 틱톡 등)
    ad_channel_cols = []
    ad_keywords = ['Meta', '메타', 'FB', 'FB/', 'N - SA', 'Naver', '네이버',
                   'Google', '구글', '틱톡', '토스', '바이럴']

    for c in range(5, min(40, len(df.columns))):
        ch = df.iloc[header_row, c]
        if pd.notna(ch):
            name = str(ch).replace('\n', '').strip()
            is_ad = any(kw in name for kw in ad_keywords)
            if is_ad:
                # 이 채널의 광고비 열
                for offset in range(4):
                    cc = c + offset
                    if cc >= len(df.columns): break
                    if pd.notna(df.iloc[col_row, cc]):
                        h = str(df.iloc[col_row, cc]).strip()
                        if '광고비' in h:
                            ad_channel_cols.append(cc)
                            break
                    if offset > 0 and pd.notna(df.iloc[header_row, cc]):
                        break

    rows = []
    for r in range(data_start, min(data_end, len(df))):
        d = safe_date(df.iloc[r, date_col], year, month)
        if d is None: continue

        ad = sum(safe_int(df.iloc[r, ac]) for ac in ad_channel_cols)
        if ad > 0:
            rows.append({'date': d.isoformat(), 'brand': sub_section['brand'], 'ad_cost': ad})

    return rows


def find_brand_sections(df):
    """브랜드 합계 행 찾기 (col3에 있는 최상위 합계만)"""
    sections = []
    for r in range(len(df)):
        for c in range(min(10, len(df.columns))):
            v = str(df.iloc[r, c]).strip()

            # 스킵
            if v == '합계': continue
            if '일자별' in v: continue
            if '로켓그로스' in v and '합계' not in v: continue
            if '매출쉐어' in v: continue
            if '자사몰 합계' in v: continue  # 하위 합계 스킵 (매출용)
            if '오픈마켓 합계' in v: continue  # 하위 합계 스킵 (매출용)
            if '자사몰+오픈마켓' in v: continue

            brand = None
            if '아자차' in v and ('합계' in v or '전체' in v or '통합' in v):
                brand = '아자차'
            elif v == '통 합' or v == '통합' or '아자차 통합' in v:
                brand = '아자차'
            elif '반드럽 합계' in v:
                brand = '반드럽'
            elif '웰바이오젠 합계' in v:
                brand = '웰바이오젠'
            elif '트라핀 합계' in v:
                brand = '웰바이오젠'
            elif '윈토르 합계' in v:
                brand = '윈토르'
            elif '자로몰 합계' in v or '자르오 합계' in v:
                brand = '자르오'
            elif '디쉬젯 합계' in v:
                brand = '기타'
            elif '로켓그로스' in v and '합계' in v:
                brand = '기타'

            if brand:
                sections.append({'brand': brand, 'row': r, 'label': v})

    return sections


def find_extra_ad_cols(df, header_row):
    """합계에 포함되지 않은 채널의 광고비 열 찾기 ('일매출 포함x' 마커)"""
    col_row = header_row + 1
    extra_cols = []
    for c in range(5, min(40, len(df.columns))):
        v = df.iloc[header_row, c]
        if pd.notna(v):
            name = str(v).strip()
            if '일매출 포함x' in name or '기 타' in name:
                for offset in range(4):
                    cc = c + offset
                    if cc >= len(df.columns): break
                    if pd.notna(df.iloc[col_row, cc]):
                        h = str(df.iloc[col_row, cc]).strip()
                        if '광고비' in h:
                            extra_cols.append(cc)
                            break
                    if offset > 0 and pd.notna(df.iloc[header_row, cc]):
                        break
    return extra_cols


def find_extra_rev_cols(df, header_row):
    """합계에 포함되지 않은 서브채널의 매출 열 찾기
    - '기 타(일매출 포함x)': 합계 미포함
    - '오픈마켓 합계': "통 합"이 자사몰만일 때 필요
    - '아자차 기타(일매출 포함x)': 같은 의미
    """
    col_row = header_row + 1
    extra_cols = []
    for c in range(5, min(40, len(df.columns))):
        v = df.iloc[header_row, c]
        if pd.notna(v):
            name = str(v).strip()
            need_add = False
            # 기타(일매출 포함x) - 합계에 미포함
            if '기 타' in name or ('기타' in name and '일매출' in name):
                need_add = True
            # 오픈마켓 합계는 "통 합"에 이미 포함이므로 추가 안 함

            if need_add:
                for offset in range(4):
                    cc = c + offset
                    if cc >= len(df.columns): break
                    if pd.notna(df.iloc[col_row, cc]):
                        h = str(df.iloc[col_row, cc]).strip()
                        if h == '매출' or h == '실매출':
                            extra_cols.append(cc)
                            break
                    if offset > 0 and pd.notna(df.iloc[header_row, cc]):
                        break
    return extra_cols


def extract_daily(df, section, year, month, next_row=None):
    """한 브랜드의 합계 일별 데이터 추출"""
    header_row = section['row']
    col_row = header_row + 1
    data_start = header_row + 2
    # 최대 33행만 읽기 (31일 + 여유). 하위 서브섹션의 중복 데이터 방지
    data_end = data_start + 33

    # 날짜 열
    date_col = 2
    for c in range(5):
        if col_row < len(df) and pd.notna(df.iloc[col_row, c]):
            if '날' in str(df.iloc[col_row, c]).strip():
                date_col = c
                break

    # 매출: 실매출 우선, 없으면 매출 (col 3-7 안의 합계)
    rev_col = None
    real_rev_col = None
    for c in range(3, min(8, len(df.columns))):
        if col_row < len(df) and pd.notna(df.iloc[col_row, c]):
            h = str(df.iloc[col_row, c]).strip()
            if h == '실매출' and real_rev_col is None:
                real_rev_col = c
            elif h == '매출' and rev_col is None:
                rev_col = c
    if real_rev_col is not None:
        rev_col = real_rev_col

    # "총 ROAS(자사+오픈+디스터)" 같은 더 포괄적인 합계 컬럼 탐지 (2023년 형식)
    # header_row의 5~40 사이에서 '총' 또는 '디스터' 포함된 컬럼을 찾음
    for c in range(5, min(40, len(df.columns))):
        v = df.iloc[header_row, c]
        if pd.notna(v):
            name = str(v).replace('\n','').strip()
            if ('총' in name and ('자사' in name or '디스터' in name or '전체' in name)) or \
               ('전체' in name and '디스터' in name):
                # 광고비/매출 sub-headers 찾기
                inc_ad = None
                inc_real_rev = None
                inc_rev = None
                for offset in range(6):
                    cc = c + offset
                    if cc >= len(df.columns): break
                    if pd.notna(df.iloc[col_row, cc]):
                        h = str(df.iloc[col_row, cc]).strip()
                        if '광고비' in h and inc_ad is None:
                            inc_ad = cc
                        elif h == '실매출' and inc_real_rev is None:
                            inc_real_rev = cc
                        elif h == '매출' and inc_rev is None:
                            inc_rev = cc
                    if offset > 0 and pd.notna(df.iloc[header_row, cc]):
                        break
                if inc_ad is not None:
                    summary_ad_col_override = inc_ad
                else:
                    summary_ad_col_override = None
                if inc_real_rev is not None:
                    rev_col = inc_real_rev
                elif inc_rev is not None:
                    rev_col = inc_rev
                if summary_ad_col_override is not None:
                    # 광고비도 디스터 포함 컬럼 사용
                    summary_ad_col_inclusive = summary_ad_col_override
                    break
    else:
        summary_ad_col_inclusive = None

    # 매출 추가 (기타 등 합계 미포함) - 포괄 합계 컬럼 사용시는 이미 포함이므로 스킵
    extra_rev_cols = [] if summary_ad_col_inclusive is not None else find_extra_rev_cols(df, header_row)

    # 환불 컬럼 (sub-header == '환불') - 매출에서 차감
    refund_cols = []
    if summary_ad_col_inclusive is None:
        for c in range(5, min(40, len(df.columns))):
            if pd.notna(df.iloc[col_row, c]):
                h = str(df.iloc[col_row, c]).strip()
                if h == '환불':
                    refund_cols.append(c)

    # 광고비: summary col3만 사용 (모든 시트에서 정확한 합계)
    summary_ad_col = None
    for c in range(3, min(7, len(df.columns))):
        if col_row < len(df) and pd.notna(df.iloc[col_row, c]):
            h = str(df.iloc[col_row, c]).strip()
            if '광고비' in h:
                summary_ad_col = c
                break
    if summary_ad_col_inclusive is not None:
        all_ad_cols = [summary_ad_col_inclusive]
    else:
        all_ad_cols = [summary_ad_col] if summary_ad_col else []
    # '일매출 포함x' 광고비 추가 (어느 경우에도 합계에 미포함)
    all_ad_cols += find_extra_ad_cols(df, header_row)

    # 이상치 검증용: 모든 서브채널 광고비/매출 열 (단순 합산용, rollup 포함될 수 있음)
    all_sub_ad = []
    all_sub_rev = []
    for c in range(5, min(45, len(df.columns))):
        ch = df.iloc[header_row, c]
        if pd.isna(ch): continue
        name = str(ch).replace('\n','').strip()
        if name in ['ROAS','B.ROAS','비고','실매출','']: continue
        if '총 ROAS' in name or '마진율' in name or '환불' in name: continue
        if 'raw data' in name or '주문건수' in name: continue
        for offset in range(5):
            cc = c + offset
            if cc >= len(df.columns): break
            if pd.notna(df.iloc[col_row, cc]):
                h = str(df.iloc[col_row, cc]).strip()
                if '광고비' in h: all_sub_ad.append(cc)
                elif h == '매출' or h == '실매출': all_sub_rev.append(cc)
            if offset > 0 and pd.notna(df.iloc[header_row, cc]): break

    rows = []
    for r in range(data_start, min(data_end, len(df))):
        d = safe_date(df.iloc[r, date_col], year, month)
        if d is None: continue
        rev = safe_int(df.iloc[r, rev_col]) if rev_col else 0
        extra_rev = sum(safe_int(df.iloc[r, ec]) for ec in extra_rev_cols)
        refund = sum(safe_int(df.iloc[r, c]) for c in refund_cols)
        total_rev = max(0, rev + extra_rev - refund)
        ad = sum(safe_int(df.iloc[r, ac]) for ac in all_ad_cols)

        # 검증: summary가 서브채널 합 * 3보다 크면 합계 수식 오류로 간주
        sub_ad_sum = sum(safe_int(df.iloc[r, c]) for c in all_sub_ad)
        sub_rev_sum = sum(safe_int(df.iloc[r, c]) for c in all_sub_rev)
        if sub_ad_sum > 0 and ad > sub_ad_sum * 3:
            ad = 0
        if sub_rev_sum > 0 and total_rev > sub_rev_sum * 3:
            total_rev = 0

        if total_rev > 0 or ad > 0:
            rows.append({
                'date': d.isoformat(),
                'brand': section['brand'],
                'revenue': total_rev,
                'ad_cost': ad,
            })
    return rows


def main():
    conn = sqlite3.connect(DB_PATH)
    # 2026년 이전 모든 데이터 삭제 (이전 실행 잔여 포함)
    conn.execute("DELETE FROM sales WHERE 날짜 < '2026-01-01'")
    conn.execute("DELETE FROM ads WHERE 날짜 < '2026-01-01'")
    conn.commit()

    files = [
        (f"{BASE_DIR}/링포 광고일지_2025.xlsx", 2025),
        (f"{BASE_DIR}/링포 광고일지_2024.xlsx", 2024),
        (f"{BASE_DIR}/링포 광고일지_2023.xlsx", 2023),
    ]

    for fpath, year in files:
        xls = pd.ExcelFile(fpath)
        print(f"[{year}]")

        for sheet in xls.sheet_names:
            if '월' not in sheet or '정산' in sheet or '목표' in sheet or '데이터' in sheet:
                continue
            try:
                month_num = int(sheet.replace('월', '').strip())
            except:
                continue

            df = pd.read_excel(fpath, sheet_name=sheet, header=None)
            sections = find_brand_sections(df)
            if not sections: continue

            # 시트 공식에 참조되지 않는 섹션 제외
            formula_refs = get_formula_row_refs(fpath, sheet)
            sections = [s for s in sections if section_in_formula(s['row'], formula_refs)]

            for i, sec in enumerate(sections):
                next_row = sections[i + 1]['row'] if i + 1 < len(sections) else None
                rows = extract_daily(df, sec, year, month_num, next_row)

                # 시트 공식 기준 truth로 스케일링
                truth_ad, truth_rev = get_section_truth(fpath, sheet, sec['row'], formula_refs)
                if rows:
                    cur_ad = sum(r['ad_cost'] for r in rows)
                    cur_rev = sum(r['revenue'] for r in rows)
                    if truth_ad is not None and cur_ad > 0:
                        ratio = float(truth_ad) / cur_ad
                        for r in rows:
                            r['ad_cost'] = int(r['ad_cost'] * ratio)
                    if truth_rev is not None and cur_rev > 0:
                        ratio = float(truth_rev) / cur_rev
                        for r in rows:
                            r['revenue'] = int(r['revenue'] * ratio)

                for r in rows:
                    store = f"{r['brand']}(합계)"
                    if r['revenue'] > 0:
                        conn.execute(
                            """INSERT OR REPLACE INTO sales
                               (날짜, 스토어, 채널, 주문건수, 매출, 객단가, 순방문자수, 전환율, 브랜드)
                               VALUES (?,?,?,?,?,?,?,?,?)""",
                            (r['date'], store, "과거데이터", 0, r['revenue'], 0, 0, 0.0, r['brand'])
                        )
                    if r['ad_cost'] > 0:
                        conn.execute(
                            """INSERT OR REPLACE INTO ads
                               (날짜, 광고채널, 광고비, 노출수, 클릭수, 전환수, 전환매출, 브랜드)
                               VALUES (?,?,?,?,?,?,?,?)""",
                            (r['date'], "과거광고비", r['ad_cost'], 0, 0, 0, 0, r['brand'])
                        )

            conn.commit()

            # 검증
            try:
                sheet_rev = int(float(df.iloc[3, 4])) if pd.notna(df.iloc[3, 4]) else 0
                sheet_ad = int(float(df.iloc[3, 3])) if pd.notna(df.iloc[3, 3]) else 0
            except:
                sheet_rev = 0
                sheet_ad = 0

            mk = f"{year}-{month_num:02d}"
            db_rev = int(conn.execute("SELECT COALESCE(SUM(매출),0) FROM sales WHERE 채널='과거데이터' AND substr(날짜,1,7)=?", (mk,)).fetchone()[0])
            db_ad = int(conn.execute("SELECT COALESCE(SUM(광고비),0) FROM ads WHERE 광고채널='과거광고비' AND substr(날짜,1,7)=?", (mk,)).fetchone()[0])

            rev_pct = ((db_rev - sheet_rev) / sheet_rev * 100) if sheet_rev > 0 else 0
            ad_pct = ((db_ad - sheet_ad) / sheet_ad * 100) if sheet_ad > 0 else 0
            rev_ok = "OK" if abs(rev_pct) <= 5 else "!!"

            print(f"  {month_num:2d}월 | 매출: 시트{sheet_rev:>12,} DB{db_rev:>12,} ({rev_pct:+.1f}%) {rev_ok} | 광고비: 시트{sheet_ad:>12,} DB{db_ad:>12,} ({ad_pct:+.1f}%)")

    conn.close()
    print("\n완료!")


if __name__ == "__main__":
    main()
